#!/usr/bin/env python3
"""
parse_import_processes.py

Parses explog_merged and extracts import process details into an Excel workbook
with two sheets: one for the import summary, one for BDoS profile configuration.

    Sheet 1 - Import Processes  : summary of every import.
    Sheet 2 - BDoS Profile Config : detailed behavioral-DoS profile parameters.

Requires:
    pip install openpyxl

Usage:
    python parse_import_processes.py
    python parse_import_processes.py --log-file ./explog_merged
    python parse_import_processes.py --log-file ./explog_merged --output ./import_processes.xlsx
"""

import argparse
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print(
        "ERROR: openpyxl is required.  Install it with:  pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)


def parse_args():
    script_dir = Path(__file__).parent
    parser = argparse.ArgumentParser(description="Import Process Log Parser")
    parser.add_argument(
        "--log-file",
        default=str(script_dir / "Input" / "explog_merged"),
        help="Path to the explog_merged file (default: Input/explog_merged)",
    )
    parser.add_argument(
        "--output",
        default=str(script_dir / "Reports" / "import_processes.xlsx"),
        help="Output Excel (.xlsx) file path (default: Reports/import_processes.xlsx)",
    )
    return parser.parse_args()


def get_param(src: str, flag: str) -> str:
    """Extract the value following *flag* from a command-line-style string."""
    m = re.search(rf"{re.escape(flag)} \"([^\"]+)\"", src)
    if m:
        return m.group(1)
    m = re.search(rf"{re.escape(flag)} (\S+)", src)
    return m.group(1) if m else ""


def write_sheet(ws, data: list, sheet_name: str) -> None:
    """Write a list of dicts to *ws* with styled headers and auto-filter."""
    ws.title = sheet_name
    if not data:
        return

    keys = list(data[0].keys())
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, key in enumerate(keys, start=1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, rec in enumerate(data, start=2):
        for col_idx, key in enumerate(keys, start=1):
            ws.cell(row=row_idx, column=col_idx, value=str(rec.get(key, "") or ""))

    ws.auto_filter.ref = ws.dimensions

    # Approximate column auto-fit
    for col_idx, key in enumerate(keys, start=1):
        max_len = len(key)
        for rec in data:
            val = str(rec.get(key, "") or "")
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def main():
    args = parse_args()
    log_file = Path(args.log_file)

    if not log_file.exists():
        print(f"ERROR: Log file not found: {log_file}", file=sys.stderr)
        sys.exit(1)

    output_xlsx = Path(args.output)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    with open(log_file, encoding="utf-8", errors="replace") as fh:
        lines = fh.readlines()

    records   = []
    bdos_recs = []

    in_block    = False
    block_lines: list[str] = []

    for line in lines:
        if re.search(r"imp_exp_audit.*\*{10,} Import process started \*{10,}", line):
            in_block    = True
            block_lines = [line]
            continue

        if not in_block:
            continue

        block_lines.append(line)

        m_end = re.search(
            r"imp_exp_audit.*\*{10,} Import process (succeeded|failed) \*{10,}", line
        )
        if not m_end:
            continue

        result = m_end.group(1)
        in_block = False

        # --- Timestamp from the "started" line ---
        timestamp = ""
        m_ts = re.search(r"\[controller (\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})", block_lines[0])
        if m_ts:
            timestamp = m_ts.group(1)

        # --- Policy name / IP from classes modify network create ---
        policy_name = ip_address = subnet = ""
        for bl in block_lines:
            if "classes modify network create" in bl:
                m = re.search(r'classes modify network create "([^"]+)"', bl)
                if m:
                    policy_name = m.group(1)
                m = re.search(r"-a ([\d.]+)", bl)
                if m:
                    ip_address = m.group(1)
                m = re.search(r"-s (\d+)", bl)
                if m:
                    subnet = m.group(1)
                break

        # --- Fields from dp policies-config table create ---
        table_key = signature = oos_profile = dos_profile = table_action = port = ""
        for bl in block_lines:
            if "dp policies-config table create" in bl:
                m = re.search(r'dp policies-config table create "([^"]+)"', bl)
                if m:
                    table_key = m.group(1)
                m = re.search(r'-sig "([^"]+)"', bl)
                if m:
                    signature = m.group(1)
                m = re.search(r'-oos "([^"]+)"', bl)
                if m:
                    oos_profile = m.group(1)
                m = re.search(r'-dos "([^"]+)"', bl)
                if m:
                    dos_profile = m.group(1)
                m = re.search(r'-a "([^"]+)"', bl)
                if m:
                    table_action = m.group(1)
                m = re.search(r"-p (\d+)", bl)
                if m:
                    port = m.group(1)
                break

        # --- Imported file name ---
        imported_file = ""
        for bl in block_lines:
            m = re.search(r"Imported file name:\s+(\S+)", bl)
            if m:
                imported_file = m.group(1)
                break

        failed_steps = sum(1 for bl in block_lines if "[STATUS: FAILED]" in bl)

        records.append({
            "Timestamp":    timestamp,
            "PolicyName":   policy_name,
            "IPAddress":    f"{ip_address}/{subnet}",
            "TableKey":     table_key,
            "Signature":    signature,
            "OOSProfile":   oos_profile,
            "BDoSProfile":  dos_profile,
            "Action":       table_action,
            "Port":         port,
            "FailedSteps":  failed_steps,
            "Result":       result,
            "ImportedFile": imported_file,
        })

        # --- BDoS profile-configuration details ---
        for bl in block_lines:
            if "dp behavioral-DoS global advanced profile-configuration create" in bl:
                bdos_recs.append({
                    "Timestamp":         timestamp,
                    "ProfileName":       policy_name,
                    "IPAddress":         f"{ip_address}/{subnet}",
                    "SYN":               get_param(bl, "-syn"),
                    "UDP":               get_param(bl, "-udp"),
                    "IGMP":              get_param(bl, "-igmp"),
                    "ICMP":              get_param(bl, "-icmp"),
                    "FragAttack":        get_param(bl, "-fa"),
                    "RST":               get_param(bl, "-rst"),
                    "SA":                get_param(bl, "-sa"),
                    "TCPFragRate":       get_param(bl, "-tfr"),
                    "UDPFragRate":       get_param(bl, "-ufr"),
                    "InboundThreshold":  get_param(bl, "-it"),
                    "OutboundThreshold": get_param(bl, "-ot"),
                    "TCPInQ":            get_param(bl, "-tiq"),
                    "UDPInQ":            get_param(bl, "-uiq"),
                    "ICMPInQ":           get_param(bl, "-iciq"),
                    "IGMPInQ":           get_param(bl, "-igiq"),
                    "UDPFragInQ":        get_param(bl, "-ufiq"),
                    "UDPFragOutQ":       get_param(bl, "-ufoq"),
                    "TCPOutQ":           get_param(bl, "-toq"),
                    "UDPOutQ":           get_param(bl, "-uoq"),
                    "ICMPOutQ":          get_param(bl, "-icoq"),
                    "IGMPOutQ":          get_param(bl, "-igoq"),
                    "Tracking":          get_param(bl, "-tr"),
                    "Profiling":         get_param(bl, "-pr"),
                    "UserSensitivity":   get_param(bl, "-usl"),
                    "Action":            get_param(bl, "-a"),
                    "Observation":       get_param(bl, "-obs"),
                    "LearnSampleTime":   get_param(bl, "-lst"),
                    "SampleType":        get_param(bl, "-st"),
                    "RateLimit":         get_param(bl, "-rl"),
                    "BurstSuppression":  get_param(bl, "-bst"),
                    "BurstInterval":     get_param(bl, "-bi"),
                })
                break

    # --- Write Excel with two sheets ---
    wb  = openpyxl.Workbook()
    ws1 = wb.active
    write_sheet(ws1, records, "Import Processes")

    ws2 = wb.create_sheet("BDoS Profile Config")
    write_sheet(ws2, bdos_recs, "BDoS Profile Config")

    if output_xlsx.exists():
        output_xlsx.unlink()
    wb.save(str(output_xlsx))

    print(f"Parsed {len(records)} import processes, {len(bdos_recs)} BDoS profile entries.")
    print(f"Output: {output_xlsx}")


if __name__ == "__main__":
    main()
